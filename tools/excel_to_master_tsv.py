# -*- coding: utf-8 -*-
"""
育成クラス出欠表_2026.xlsx の「出欠管理票_2026」シートから、
生徒マスタ向け TSV（studentId, 氏名, コース, レギュラー曜日, 初期残振替, メモ）を生成する。
"""
from __future__ import annotations

import os
import re
import sys

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

SHEET_NAME = "出欠管理票_2026"


def is_date_cell(val) -> bool:
    if val is None:
        return False
    if hasattr(val, "year") and hasattr(val, "month"):
        return True
    s = str(val)
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}", s))


def weekday_token(s: str) -> str | None:
    s = str(s).strip()
    if not s or s in ("", "休会"):
        return None
    for d in ("月", "火", "水", "木", "金", "土", "日"):
        if d in s:
            return d
    return None


def main() -> int:
    default_xlsx = os.path.join(os.environ.get("USERPROFILE", "."), "Downloads", "育成クラス出欠表_2026.xlsx")
    path = sys.argv[1] if len(sys.argv) > 1 else default_xlsx
    if not os.path.isfile(path):
        print("ファイルが見つかりません:", path, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print("シートがありません:", SHEET_NAME, "/ あるシート:", wb.sheetnames, file=sys.stderr)
        wb.close()
        return 1

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row_idx = None
    first_date_col = None
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        a0 = row[0] if len(row) > 0 else None
        if a0 and "名前" in str(a0) and "日付" in str(a0):
            header_row_idx = i
            for j, cell in enumerate(row):
                if is_date_cell(cell):
                    first_date_col = j
                    break
            break

    if header_row_idx is None or first_date_col is None:
        print("見出し行または日付列を特定できませんでした。", file=sys.stderr)
        return 1

    out_lines = []
    out_lines.append("studentId\t氏名\tコース\tレギュラー曜日\t初期残振替\tメモ")
    n = 1
    for row in rows[header_row_idx + 1 :]:
        if not row or len(row) < 2:
            continue
        name = row[0]
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        grade = row[1] if len(row) > 1 else ""
        grade = "" if grade is None else str(grade).strip()

        days: list[str] = []
        for j in range(2, min(first_date_col, len(row))):
            cell = row[j]
            if cell is None or str(cell).strip() == "":
                continue
            w = weekday_token(cell)
            if w and w not in days:
                days.append(w)

        if not days:
            continue

        week_n = len(days)
        if week_n >= 3:
            course = "週3"
        elif week_n == 2:
            course = "週2"
        else:
            course = "週1"

        reg = ",".join(days)
        sid = "S" + str(n).zfill(3)
        memo = "学年:" + grade if grade else ""
        out_lines.append("\t".join([sid, name, course, reg, "0", memo]))
        n += 1

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "master_from_出欠管理票_2026.tsv")
    with open(out_path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(out_lines) + "\n")

    print("出力:", out_path)
    print("行数:", len(out_lines) - 1)
    return 0


if __name__ == "__main__":
    sys.exit(main())
