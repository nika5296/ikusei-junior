# -*- coding: utf-8 -*-
"""育成クラス出欠表_2026.xlsx のシート名と先頭行を表示"""
import os
import sys

path = os.path.join(os.environ.get("USERPROFILE", ""), "Downloads", "育成クラス出欠表_2026.xlsx")
if len(sys.argv) > 1:
    path = sys.argv[1]

if not os.path.isfile(path):
    print("FILE_NOT_FOUND:", path)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("FILE:", path)
print("SHEETS:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
        rows.append(row)
        if i >= 24:
            break
    print("\n===", name, "===")
    for r in rows:
        line = []
        for c in r[:20]:
            if c is None:
                line.append("")
            else:
                s = str(c).replace("\n", " ")
                if len(s) > 40:
                    s = s[:37] + "..."
                line.append(s)
        print(" | ".join(line))
wb.close()
