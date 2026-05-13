# -*- coding: utf-8 -*-
"""育成クラス出欠管理用 Excel テンプレート（Googleスプレッドシートへ取り込み可）"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT_NAME = "育成クラス出欠_テンプレート.xlsx"
OUT_IMPORT_PRESET = "育成クラス出欠_インポート用_フォーム設定プリセット.xlsx"
MASTER_TSV = "data/生徒マスタ_2026.tsv"


def build_settings_rows_preset():
    """Setup.gs buildSettingsTemplate_ と同一キー。横持ち・今日の日付・出席のみフォーム向けの値。"""
    return [
        ["キー", "値", "説明"],
        [
            "予約ページURL",
            "https://reserva.be/kamakuragreen",
            "保護者向け「振替予約」ボタンの遷移先（必要に応じて変更）",
        ],
        ["アラート送信先メール", "vnika0305@gmail.com", "未入力アラートの送信先（変更可）"],
        [
            "出欠フォームURL",
            "https://docs.google.com/forms/d/e/1FAIpQLSeW7H0GdT9OZYmfhxmc_5AgsdyKhCIcAJuNy9SVdu7GWVbIsQ/viewform",
            "未入力アラート本文に載せる育成クラス出欠フォームのURL",
        ],
        ["タイムゾーン", "Asia/Tokyo", "日付・曜日判定の基準"],
        ["未入力アラート除外日", "", "カンマ区切り yyyy-MM-dd。学校都合で実施なしの日はここに入れると前日アラートを抑止"],
        [
            "フォーム_実施日列名",
            "今日の日付",
            "回答シート1行目の「実施日」列見出しと完全一致（Googleフォームの設問タイトルに合わせる）",
        ],
        [
            "フォーム_所属曜日列名",
            "所属曜日",
            "所属曜日列を付けていない場合は見出しは存在しなくてよい（下の「実施日から」で暦曜日判定）",
        ],
        [
            "フォーム_所属曜日なしは実施日から",
            "TRUE",
            "TRUE: 所属曜日列が無くても、実施日の暦曜日でクラスを決める",
        ],
        [
            "フォーム_出席のみで空欄は欠席",
            "TRUE",
            "TRUE: 出席にチェックした列だけ「出席」が入り、空欄は欠席として出欠ログに残す",
        ],
        ["フォーム_コーチ名列名", "コーチ名", "任意。フォームに列が無ければ値は未使用で問題なし"],
        [
            "Surge_スタイルシートURL",
            "https://kamakura-green-ikusei.surge.sh/styles.css",
            "Webアプリ用。surge の styles.css を公開したURL（Config.gs の既定と合わせる）",
        ],
    ]


def build_import_instruction_rows():
    """Googleスプレッドシートへの取り込み手順（このシートは GAS では未使用）。"""
    return [
        ["項目", "内容"],
        [
            "このファイルの目的",
            "「設定」シートのキー・値を、スプレッドシートの「設定」に反映するためのプリセットです。",
        ],
        [
            "手順①",
            "Googleスプレッドシートで「設定」シートを開き、キー列（A列）に同じキーがある行は B列の値を、このファイルの「設定」と同じに書き換える。",
        ],
        [
            "手順②（まとめて貼り付け）",
            "「設定」シートの A2:C 付近を選択し、この Excel の「設定」から 2 行目以降をコピーして貼り付けてもよい（1行目の見出し「キー・値・説明」は重複しないよう調整）。",
        ],
        [
            "フォーム回答シート名",
            "回答先は「フォーム回答」または Google 標準の「フォームの回答」のどちらでもスクリプトが認識します。",
        ],
        [
            "生徒列の見出し",
            "「生徒名 [氏名]」形式は可。生徒マスタの「氏名」と一致する必要があります（空白の全角/半角の違いはある程度吸収されます）。",
        ],
        [
            "集計・出欠ログ",
            "手入力不要。フォーム送信後にスクリプトが追記・再計算します。",
        ],
    ]


def build_form_header_example_rows():
    """回答シート1行目のイメージ（実際のフォーム連携後はフォームが上書きします）。"""
    return [
        [
            "※下の1行は見本です。実際の列数・氏名はフォームに合わせてください。",
        ],
        [
            "タイムスタンプ",
            "今日の日付",
            "生徒名 [例・西百花]",
            "生徒名 [例・高橋陽仁]",
            "（以下、生徒ごとに列が続く）",
        ],
    ]


def load_master_rows(root: Path) -> list[list]:
    """data/生徒マスタ_2026.tsv があれば読み込み、なければサンプル1行"""
    p = root / MASTER_TSV
    if not p.exists():
        return [
            [
                "studentId",
                "氏名",
                "コース",
                "レギュラー曜日",
                "初期残振替",
                "viewToken",
                "token無効",
                "メモ",
            ],
            [
                "S001",
                "サンプル太郎",
                "週2",
                "月,木",
                "0",
                "",
                "",
                "初期行は削除してOK。コースは週1/週2/週3。レギュラー曜日はカンマ区切り（月,木）",
            ],
        ]
    lines = p.read_text(encoding="utf-8").strip().splitlines()
    rows = []
    for line in lines:
        rows.append(line.split("\t"))
    return rows


def autosize_cols(ws, max_col=None, min_width=10, max_width=52):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    L = len(str(cell.value))
                    if L > best:
                        best = min(L + 2, max_width)
        ws.column_dimensions[letter].width = best


NO_BOLD_HEADER = frozenset({"フォーム回答", "既存表取込_作業", "フォーム見出し例"})


def write_workbook(path: Path, sheets_data: list[tuple[str, list]]) -> None:
    wb = Workbook()
    wb.active.title = sheets_data[0][0]
    for name, _rows in sheets_data[1:]:
        wb.create_sheet(title=name)

    for (name, rows), sheet in zip(sheets_data, wb.worksheets):
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                cell = sheet.cell(row=r, column=c, value=val)
                bold = r == 1 and name not in NO_BOLD_HEADER
                if name == "フォーム見出し例" and r == 2:
                    bold = True
                if bold:
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize_cols(sheet)

    wb.save(path)
    print(f"Wrote: {path}")


def main():
    root = Path(__file__).resolve().parent.parent
    master_rows = load_master_rows(root)

    common_front = [
        ("設定", build_settings_rows_preset()),
        ("インポート手順", build_import_instruction_rows()),
        ("フォーム見出し例", build_form_header_example_rows()),
    ]

    sheets_full = [
        *common_front,
        ("生徒マスタ", master_rows),
        (
            "フォーム回答",
            [["※Googleフォームをこのスプレッドシートに接続すると、1行目に質問見出しが入ります"]],
        ),
        (
            "出欠ログ",
            [
                [
                    "logId",
                    "studentId",
                    "実施日",
                    "所属曜日",
                    "出欠",
                    "欠席による振替発生",
                    "振替消化",
                    "元フォーム行",
                    "フォーム受信時刻",
                    "メモ",
                ],
            ],
        ),
        (
            "集計",
            [["studentId", "氏名", "振替発生回数", "振替消化回数", "残振替回数", "最終集計日時"]],
        ),
        ("通知管理", [["送信日時", "種別", "内容", "結果"]]),
        (
            "既存表取込_作業",
            [
                ["※既存「出欠管理票_2026」の4月表から、氏名・曜日・コースをコピペ（下の見出し行は残す）"],
                ["氏名", "レギュラー曜日（カンマ区切り: 月,木）", "コース（週1/週2/週3）"],
            ],
        ),
    ]

    write_workbook(root / OUT_NAME, sheets_full)

    sheets_preset_only = list(common_front)
    write_workbook(root / OUT_IMPORT_PRESET, sheets_preset_only)


if __name__ == "__main__":
    main()
